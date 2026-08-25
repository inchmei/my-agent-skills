"""Apply visual styles to Excel worksheets based on template config."""

import math
from copy import copy
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers, Color
)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def _rgb_color(hex_str):
    """Create an openpyxl Color from a hex string (e.g. 'FFFFFF' or '#FFFFFF').

    6-digit hex gets an 'FF' alpha prefix (fully opaque) — a '00' prefix
    (fully transparent) breaks rendering in some Excel/WPS versions.
    """
    h = hex_str.lstrip("#").upper()
    if len(h) == 6:
        return Color(rgb=f"FF{h}")
    return Color(rgb=h)


def resolve_color(template, color_ref):
    """Resolve 'palette.primary' references to hex strings."""
    if not color_ref:
        return None
    if not isinstance(color_ref, str):
        return color_ref
    if color_ref.startswith("palette."):
        key = color_ref.split(".", 1)[1]
        return template["global"]["palette"].get(key, "000000")
    return color_ref


def hex_to_rgb(h):
    """Convert hex string to (r, g, b) tuple."""
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))


def reset_cell_styles(ws, structure):
    """Strip all pre-existing cell styles so theme rules apply uniformly.

    Called first in apply_all: resets font / fill / border / alignment /
    number_format back to defaults for every cell in the table region
    (row 1 through the last data row), and clears any pre-existing
    conditional formatting rules. Values, formulas and merged cells are
    untouched. Cells without value and without explicit style are skipped
    to keep the output file lean.
    """
    cs = structure["col_start"]
    ce = structure["col_end"]
    de = structure["data_end"]
    for row in ws.iter_rows(min_row=1, max_row=de, min_col=cs, max_col=ce):
        for cell in row:
            if cell.value is None and not cell.has_style:
                continue
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()
            cell.number_format = "General"
    try:
        from openpyxl.formatting.formatting import ConditionalFormattingList
        ws.conditional_formatting = ConditionalFormattingList()
    except Exception:
        pass  # some openpyxl versions expose conditional_formatting as list


def apply_global_font(ws, template, structure):
    """Apply base font to all data rows."""
    font_name = template["global"]["font"]
    base_size = template["global"]["base_size"]
    data_font = Font(name=font_name, size=base_size)

    cs = structure["col_start"]
    ce = structure["col_end"]
    for row in range(structure["data_start"], structure["data_end"] + 1):
        for col in range(cs, ce + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font


def apply_title(ws, template, structure):
    """Apply title row styling."""
    title_cfg = template.get("title", {})
    if not title_cfg.get("enabled") or not structure["title_row"]:
        return

    row = structure["title_row"]
    if row < 1:
        return

    cs = structure["col_start"]
    ce = structure["col_end"]
    palette = template["global"]["palette"]

    bg = resolve_color(template, title_cfg["bg"])
    color = title_cfg.get("color", "FFFFFF")
    font_name = template["global"]["font"]
    font_size = title_cfg.get("size", 14)
    align = title_cfg.get("align", "center")
    height = title_cfg.get("height", 30)

    title_font = Font(name=font_name, size=font_size, bold=True, color=_rgb_color(color))
    title_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    title_align = Alignment(horizontal=align, vertical="center", wrap_text=True)

    # Merge columns if enabled (skip if the row already has merged cells,
    # since openpyxl raises on overlapping merge ranges)
    if title_cfg.get("merge_columns") and cs < ce:
        has_existing_merge = any(
            m.min_row <= row <= m.max_row for m in ws.merged_cells.ranges
        )
        if not has_existing_merge:
            ws.merge_cells(
                start_row=row, start_column=cs, end_row=row, end_column=ce
            )

    for col in range(cs, ce + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_align

    ws.row_dimensions[row].height = height


def apply_header(ws, template, structure):
    """Apply header row styling."""
    header_cfg = template.get("header", {})
    row = structure["header_row"]
    cs = structure["col_start"]
    ce = structure["col_end"]

    bg = resolve_color(template, header_cfg["bg"])
    color = header_cfg.get("color", "FFFFFF")
    font_name = template["global"]["font"]
    bold = header_cfg.get("bold", True)
    align = header_cfg.get("align", "center")
    height = header_cfg.get("height", 24)

    header_font = Font(name=font_name, size=10, bold=bold, color=_rgb_color(color))
    header_fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    header_align = Alignment(horizontal=align, vertical="center", wrap_text=True)

    for col in range(cs, ce + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[row].height = height


def apply_body(ws, template, structure):
    """Apply body area styling: zebra rows, borders, alignment, font."""
    body_cfg = template.get("body", {})
    ds = structure["data_start"]
    de = structure["data_end"]
    cs = structure["col_start"]
    ce = structure["col_end"]

    font_name = template["global"]["font"]
    font_size = body_cfg.get("size", 10)
    align = body_cfg.get("align", "left")
    body_fill_color = resolve_color(template, body_cfg.get("body_fill"))
    zebra_color = resolve_color(template, body_cfg.get("zebra"))
    border_cfg = body_cfg.get("border", {})

    body_wrap = body_cfg.get("wrap_text", False)
    body_font = Font(name=font_name, size=font_size)
    body_align = Alignment(horizontal=align, vertical="center", wrap_text=body_wrap)

    thin_side = Side(style=border_cfg.get("style", "thin"), color=border_cfg.get("color", "BFBFBF"))
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    body_fill = PatternFill(start_color=body_fill_color, end_color=body_fill_color, fill_type="solid") if body_fill_color else None
    zebra_fill = PatternFill(start_color=zebra_color, end_color=zebra_color, fill_type="solid") if zebra_color else None

    for row_idx, row in enumerate(range(ds, de + 1)):
        for col in range(cs, ce + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = border

            # body_fill applies to all rows; zebra overrides on odd rows if set
            if zebra_fill and row_idx % 2 == 1:
                cell.fill = zebra_fill
            elif body_fill:
                cell.fill = body_fill


def apply_number_formats(ws, template, structure, col_types):
    """Apply number formatting based on detected column types."""
    format_cfg = template.get("formats", {})
    ds = structure["data_start"]
    de = structure["data_end"]

    type_to_format = {}
    for fmt_name, cfg in format_cfg.items():
        detect_type = cfg.get("detect", "")
        if detect_type:
            type_to_format[detect_type] = cfg.get("pattern", "")

    for col, col_type in col_types.items():
        fmt = type_to_format.get(col_type)
        if fmt:
            for row in range(ds, de + 1):
                cell = ws.cell(row=row, column=col)
                cell.number_format = fmt


def normalize_percent_values(ws, structure, col_types):
    """Convert string percentages (e.g. '12.3%') to numeric values (0.123).

    This is required for conditional formatting (CellIsRule >0 / <0) to work,
    since Excel cannot compare strings against numeric thresholds.
    """
    ds = structure["data_start"]
    de = structure["data_end"]

    for col, ctype in col_types.items():
        if ctype != "percent":
            continue
        for row in range(ds, de + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if isinstance(val, str) and val.strip().endswith("%"):
                try:
                    num = float(val.strip().rstrip("%")) / 100
                    cell.value = num
                except ValueError:
                    pass


def apply_conditional_formatting(ws, template, structure, col_types):
    """Apply conditional formatting for up/down colors.

    Supports:
    - column_types: limit to specific types (default: number, integer, percent)
    - column_header_keywords: only apply to columns whose header contains any keyword
    - configurable font_color and bold per up/down rule
    """
    cond_cfg = template.get("conditional", {})
    if not cond_cfg:
        return

    up_color = cond_cfg["up"]["color"]
    down_color = cond_cfg["down"]["color"]
    up_font_color = cond_cfg["up"].get("font_color", "FFFFFF")
    down_font_color = cond_cfg["down"].get("font_color", "FFFFFF")
    up_bold = cond_cfg["up"].get("bold", True)
    down_bold = cond_cfg["down"].get("bold", True)

    # Which column types to target (default: all numeric)
    target_types = cond_cfg.get("column_types", ["number", "integer", "percent"])
    # Optional: only columns whose header contains any of these keywords
    target_keywords = cond_cfg.get("column_header_keywords", [])

    ds = structure["data_start"]
    de = structure["data_end"]
    header_row = structure["header_row"]

    # Find matching columns
    numeric_cols = []
    for col, ctype in col_types.items():
        if ctype not in target_types:
            continue
        if target_keywords:
            header_val = str(ws.cell(row=header_row, column=col).value or "")
            if not any(kw in header_val for kw in target_keywords):
                continue
        numeric_cols.append(col)

    for col in numeric_cols:
        col_letter = get_column_letter(col)
        cell_range = f"{col_letter}{ds}:{col_letter}{de}"

        up_fill = PatternFill(start_color=up_color, end_color=up_color, fill_type="solid")
        up_font = Font(color=_rgb_color(up_font_color), bold=up_bold)
        down_fill = PatternFill(start_color=down_color, end_color=down_color, fill_type="solid")
        down_font = Font(color=_rgb_color(down_font_color), bold=down_bold)

        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="greaterThan", formula=["0"], fill=up_fill, font=up_font),
        )
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator="lessThan", formula=["0"], fill=down_fill, font=down_font),
        )


# Columns whose header contains these words are treated as descriptive text,
# never as a level/status semantic column (e.g. "风险描述", "状态说明").
_SEMANTIC_EXCLUDE_HEADERS = ("描述", "说明", "详情", "备注", "内容", "举证", "建议", "措施")


def _is_light_color(hex_str):
    """Return True if a hex color is light enough that white text would be unreadable."""
    if not hex_str:
        return False
    try:
        r, g, b = hex_to_rgb(hex_str)
    except (ValueError, TypeError):
        return False
    return (0.299 * r + 0.587 * g + 0.114 * b) > 160


def _match_semantic(text, mapping):
    """Match cell text against a semantic color mapping.

    Exact match wins, then longest-key containment (so "中低危" matches before
    the single-char aliases "中" / "低" do).
    """
    t = str(text).strip()
    if not t:
        return None
    for key in sorted(mapping, key=len, reverse=True):
        if t == key or key in t:
            return mapping[key]
    return None


def _find_semantic_cols(ws, header_row, cs, ce, keywords):
    """Find columns whose header contains any keyword (excluding descriptive headers)."""
    cols = []
    for col in range(cs, ce + 1):
        header = str(ws.cell(row=header_row, column=col).value or "").strip()
        if not header:
            continue
        if any(x in header for x in _SEMANTIC_EXCLUDE_HEADERS):
            continue
        if any(kw in header for kw in keywords):
            cols.append(col)
    return cols


def apply_semantic_colors(ws, template, structure):
    """Color level / status / result columns based on the template's semantic maps.

    Each group in template["semantic"]["columns"] maps header keywords to a
    color map named "<group>_map" (e.g. level -> level_map, status -> status_map,
    result -> result_map). For every data cell in matched columns, the cell text
    is matched against the group's map and the mapped color is applied. Style is
    "fill" (solid fill + font color) or "text" (font color only, keep background).
    Light mapped colors automatically switch to dark text for readability.
    """
    sem_cfg = template.get("semantic", {})
    if not sem_cfg.get("enabled", True):
        return

    col_cfg = sem_cfg.get("columns", {})
    if not col_cfg:
        return

    style = sem_cfg.get("style", "fill")
    font_color = sem_cfg.get("font_color", "FFFFFF")
    bold = sem_cfg.get("bold", True)

    cs = structure["col_start"]
    ce = structure["col_end"]
    header_row = structure["header_row"]
    ds = structure["data_start"]
    de = structure["data_end"]

    # Each configured column group -> its own <group>_map (may be absent, skip).
    targets = []
    for group, cfg in col_cfg.items():
        mapping = sem_cfg.get(f"{group}_map", {})
        if not mapping:
            continue
        keywords = cfg.get("keywords", [])
        cols = _find_semantic_cols(ws, header_row, cs, ce, keywords)
        if cols:
            targets.append((cols, mapping))

    for cols, mapping in targets:
        for col in cols:
            for row in range(ds, de + 1):
                cell = ws.cell(row=row, column=col)
                color = _match_semantic(cell.value, mapping)
                if not color:
                    continue

                new_font = copy(cell.font)
                if style == "text":
                    new_font.color = _rgb_color(color)
                    new_font.bold = bold
                    cell.font = new_font
                else:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    # Light fills (e.g. #FDAA1D amber) need dark text to stay readable
                    text_color = "2F3540" if _is_light_color(color) else font_color
                    new_font.color = _rgb_color(text_color)
                    new_font.bold = bold
                    cell.font = new_font


def apply_column_styles(ws, template, structure):
    """Apply per-column style overrides matched by header keywords.

    Each spec in template["column_styles"]:
      { "keywords": [...], "bold"?: bool, "color"?: hex, "fill"?: hex|null }
    - bold / color update the font (missing keys keep the current value)
    - fill hex sets a solid fill; fill null clears the fill (missing key keeps it)
    Runs after semantic colors so it cannot be reverted by them.
    """
    specs = template.get("column_styles") or []
    if not specs:
        return

    header_row = structure["header_row"]
    ds = structure["data_start"]
    de = structure["data_end"]
    cs = structure["col_start"]
    ce = structure["col_end"]

    for spec in specs:
        kws = spec.get("keywords") or []
        if not kws:
            continue
        cols = [col for col in range(cs, ce + 1)
                if any(kw in str(ws.cell(row=header_row, column=col).value or "") for kw in kws)]
        for col in cols:
            for row in range(ds, de + 1):
                cell = ws.cell(row=row, column=col)
                f = cell.font
                if "bold" in spec or spec.get("color"):
                    new_bold = spec["bold"] if "bold" in spec else f.bold
                    new_color = _rgb_color(spec["color"]) if spec.get("color") else f.color
                    cell.font = Font(name=f.name, size=f.size, bold=new_bold,
                                     italic=f.italic, underline=f.underline, color=new_color)
                if "fill" in spec:
                    if spec["fill"]:
                        cell.fill = PatternFill(start_color=spec["fill"], end_color=spec["fill"], fill_type="solid")
                    else:
                        cell.fill = PatternFill(fill_type=None)


def apply_page_setup(ws, template, structure):
    """Apply page layout settings."""
    page_cfg = template.get("page", {})
    from openpyxl.worksheet.page import PageMargins

    orientation = page_cfg.get("orientation", "landscape")
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1 if page_cfg.get("fit_to_width") else 0
    ws.page_setup.fitToHeight = 0

    margins = page_cfg.get("margins", {})
    ws.page_margins = PageMargins(
        top=margins.get("top", 0.5),
        bottom=margins.get("bottom", 0.5),
        left=margins.get("left", 0.5),
        right=margins.get("right", 0.5),
        header=0.3,
        footer=0.3,
    )


def auto_fit_column_widths(ws, structure):
    """Auto-fit column widths based on content."""
    cs = structure["col_start"]
    ce = structure["col_end"]
    de = structure["data_end"]

    for col in range(cs, ce + 1):
        max_width = 8
        for row in range(structure["header_row"], de + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # Approximate: CJK chars ~2x width, ASCII ~1x
                width = 0
                for ch in str(val):
                    width += 2 if ord(ch) > 127 else 1
                max_width = max(max_width, width)
        ws.column_dimensions[get_column_letter(col)].width = min(max_width + 4, 40)


def estimate_row_heights(ws, template, structure):
    """Set body row heights to max(min_height, estimated content height).

    Excel has no native "min height + auto-grow" mode — an explicit height
    is fixed, no height means fully auto. So we estimate each row's needed
    height from its content (font size, column widths, wrap_text) and set
    the row height to max(min_height, estimated). Long content rows grow,
    short rows keep the minimum.
    """
    body_cfg = template.get("body", {})
    min_height = body_cfg.get("min_height")
    if not min_height:
        return

    font_size = body_cfg.get("size", 10)
    wrap = body_cfg.get("wrap_text", False)
    max_height = body_cfg.get("max_height")  # optional hard cap (points)
    ds = structure["data_start"]
    de = structure["data_end"]
    cs = structure["col_start"]
    ce = structure["col_end"]

    # Estimated height of one wrapped line: font size * line factor + padding
    line_height = font_size * 1.35 + 2

    # A single text segment needing more lines than this would never fit the
    # visible row anyway (e.g. a 20k-char "evidence" field). Skip such columns
    # entirely so they cannot blow up the row height estimate.
    long_text_line_cap = body_cfg.get("long_text_line_cap", 8)

    for row in range(ds, de + 1):
        max_lines = 1
        for col in range(cs, ce + 1):
            cell = ws.cell(row=row, column=col)
            val = cell.value
            if val is None:
                continue
            s = str(val).strip()
            if not s:
                continue

            # Split by explicit newlines first: each segment is its own
            # display line and wraps independently (e.g. "url: x\nbody: y"
            # renders as 2+ lines even if each segment is short).
            segments = s.split("\n")
            cell_lines = 0
            if wrap:
                col_width = ws.column_dimensions[get_column_letter(col)].width or 8
                usable = max(col_width - 1, 1)
                for seg in segments:
                    seg = seg.strip()
                    if not seg:
                        cell_lines += 1
                        continue
                    # Content width in ASCII-char units (CJK ~2x)
                    seg_width = sum(2 if ord(ch) > 127 else 1 for ch in seg)
                    seg_lines = max(1, math.ceil(seg_width / usable))
                    if seg_lines > long_text_line_cap:
                        # Unreasonably long field — skip this column entirely
                        cell_lines = 0
                        break
                    cell_lines += seg_lines
            else:
                cell_lines = max(1, len(segments))
            max_lines = max(max_lines, cell_lines)

        height = max(min_height, max_lines * line_height)
        if max_height:
            height = min(height, max_height)
        ws.row_dimensions[row].height = height


def apply_all(ws, template, structure, col_types):
    """Apply all visual styles to a worksheet."""
    reset_cell_styles(ws, structure)
    apply_global_font(ws, template, structure)
    apply_title(ws, template, structure)
    apply_header(ws, template, structure)
    apply_body(ws, template, structure)
    normalize_percent_values(ws, structure, col_types)
    apply_number_formats(ws, template, structure, col_types)
    apply_conditional_formatting(ws, template, structure, col_types)
    apply_semantic_colors(ws, template, structure)
    apply_column_styles(ws, template, structure)
    apply_page_setup(ws, template, structure)
    auto_fit_column_widths(ws, structure)
    estimate_row_heights(ws, template, structure)
