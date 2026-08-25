"""Auto-detect table structure: title row, header row, data range, column types."""

import re
from openpyxl.utils import get_column_letter


def detect_structure(ws):
    """
    Detect the structure of a worksheet.
    Returns: {
        "title_row": int or None,
        "header_row": int,
        "data_start": int,
        "data_end": int,
        "col_start": int,
        "col_end": int
    }
    """
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Step 1: Detect header row (first row where most columns have content)
    header_row = _detect_header_row(ws, max_row, max_col)
    if header_row is None:
        # No reliable header found — treat as non-table sheet (e.g. cover/notes page)
        return None

    # Step 2: Detect title row (row before header with distinctive formatting)
    title_row = _detect_title_row(ws, header_row, max_col)

    # Step 3: Detect data range
    data_start = header_row + 1
    data_end = _detect_data_end(ws, data_start, max_row, max_col)
    col_start, col_end = _detect_column_range(ws, data_start, data_end, max_col)

    return {
        "title_row": title_row,
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "col_start": col_start,
        "col_end": col_end,
    }


def _detect_header_row(ws, max_row, max_col):
    """Find the first row where most non-empty columns have values.

    Returns the row index, or None if no reliable header row exists
    in the first few rows (e.g. a cover/notes sheet, not a table).
    """
    search_rows = min(max_row, 5)
    best_row, best_score = None, 0

    for row in range(1, search_rows + 1):
        filled = 0
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                filled += 1
        if filled > best_score and filled >= max_col * 0.5:
            best_score = filled
            best_row = row

    return best_row


def _detect_title_row(ws, header_row, max_col):
    """Check if there's a title row above the header."""
    if header_row <= 1:
        return None

    title_candidate = header_row - 1

    # If no content at all, skip
    has_content = False
    for col in range(1, max_col + 1):
        if ws.cell(row=title_candidate, column=col).value is not None:
            has_content = True
            break
    if not has_content:
        return None

    # Check for merged cells on this row
    for merge in ws.merged_cells.ranges:
        if merge.min_row == title_candidate:
            return title_candidate

    # Row 1 with content, header is row 2 — highly likely a title
    if title_candidate == 1:
        return title_candidate

    # Check if first cell has larger font or colored background
    cell = ws.cell(row=title_candidate, column=1)
    if cell.font and cell.font.size and cell.font.size >= 12:
        return title_candidate
    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb and cell.fill.fgColor.rgb != "00000000":
        return title_candidate

    return None


def _detect_data_end(ws, data_start, max_row, max_col):
    """Find the last data row (stop at first completely empty row)."""
    consecutive_empty = 0
    for row in range(data_start, max_row + 1):
        has_content = False
        for col in range(1, max_col + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_content = True
                break
        if has_content:
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            if consecutive_empty >= 2:
                return row - 2
    return max_row


def _detect_column_range(ws, data_start, data_end, max_col):
    """Find the leftmost and rightmost columns with data."""
    col_start, col_end = 1, max_col

    for col in range(1, max_col + 1):
        has_data = False
        for row in range(data_start, data_end + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            col_start = col
            break

    for col in range(max_col, 0, -1):
        has_data = False
        for row in range(data_start, data_end + 1):
            if ws.cell(row=row, column=col).value is not None:
                has_data = True
                break
        if has_data:
            col_end = col
            break

    return col_start, col_end


def detect_column_types(ws, header_row, data_start, data_end, col_start, col_end):
    """
    Classify each column by data type.
    Returns: dict {col_index: "number"|"percent"|"date"|"integer"|"text"}
    """
    types = {}
    for col in range(col_start, col_end + 1):
        samples = []
        for row in range(data_start, min(data_start + 20, data_end + 1)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                samples.append(str(val))

        if not samples:
            types[col] = "text"
            continue

        types[col] = _classify_samples(samples)
    return types


def _classify_samples(samples):
    """Classify a column based on its sample values."""
    n = len(samples)

    # Check percent: ends with %
    percent_count = sum(1 for s in samples if s.strip().endswith("%"))
    if percent_count >= n * 0.5:
        return "percent"

    # Check date patterns
    date_count = sum(1 for s in samples if _looks_like_date(s))
    if date_count >= n * 0.5:
        return "date"

    pure_numbers = [_to_float(s) for s in samples]
    pure_numbers = [v for v in pure_numbers if v is not None]

    if len(pure_numbers) < n * 0.5:
        return "text"

    # Distinguish integer vs decimal
    int_count = sum(1 for v in pure_numbers if v == int(v))
    if int_count >= len(pure_numbers) * 0.8:
        return "integer"

    return "number"


def _looks_like_date(s):
    """Check if a string looks like a date."""
    s = s.strip()
    patterns = [
        r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}$",
        r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}$",
        r"^\d{4}年\d{1,2}月\d{1,2}[日号]?$",
        r"^\d{1,2}月\d{1,2}[日号]$",
    ]
    return any(re.match(p, s) for p in patterns)


def _to_float(s):
    """Convert a string to float, removing thousand separators. Returns None on failure."""
    try:
        cleaned = s.strip().replace(",", "").replace("¥", "").replace("$", "").replace(" ", "")
        # Handle parentheses for negative numbers
        if cleaned.startswith("(") and cleaned.endswith(")"):
            cleaned = "-" + cleaned[1:-1]
        return float(cleaned)
    except ValueError:
        return None
