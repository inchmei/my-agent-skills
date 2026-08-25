"""Core pipeline: load template, detect structure, apply styles, save output."""

import csv
import io
import json
from pathlib import Path
from openpyxl import Workbook, load_workbook

from .detector import detect_structure, detect_column_types
from .styles import apply_all
from .watermark import apply_watermark


# Built-in themes directory
THEMES_DIR = Path(__file__).parent / "themes"


def load_template(theme="business_blue", custom_path=None):
    """Load a template by name (built-in theme) or from custom path."""
    if custom_path:
        path = Path(custom_path)
    else:
        path = THEMES_DIR / f"{theme}.json"
        if not path.exists():
            raise FileNotFoundError(
                f"Theme '{theme}' not found at {path}. "
                f"Available themes: {_list_themes()}"
            )

    with open(path, "r", encoding="utf-8") as f:
        template = json.load(f)

    _resolve_refs(template, path.parent)

    # Validate required sections
    _validate_template(template)
    return template


def _resolve_refs(template, base_dir):
    """Resolve '$ref' keys in template sections by loading shared JSON files.

    Theme files can reference shared configuration, e.g.
    "semantic": {"$ref": "shared_semantic.json"} — the referenced file's
    top-level object replaces the whole section. This keeps shared blocks
    (semantic color maps) defined once instead of duplicated per theme.
    """
    for key, value in list(template.items()):
        if isinstance(value, dict) and isinstance(value.get("$ref"), str):
            ref_path = (base_dir / value["$ref"]).resolve()
            with open(ref_path, "r", encoding="utf-8") as f:
                shared = json.load(f)
            # Drop internal bookkeeping keys (e.g. _comment)
            shared.pop("_comment", None)
            template[key] = shared
        elif isinstance(value, dict):
            _resolve_refs(value, base_dir)


def _list_themes():
    """List available built-in theme names."""
    if not THEMES_DIR.exists():
        return []
    return sorted([p.stem for p in THEMES_DIR.glob("*.json")])


def _validate_template(template):
    """Basic validation of template structure."""
    required_sections = ["global", "title", "header", "body"]
    for section in required_sections:
        if section not in template:
            raise ValueError(f"Template missing required section: '{section}'")

    if "palette" not in template["global"]:
        raise ValueError("Template missing 'global.palette'")

    required_colors = ["primary", "accent", "muted", "zebra"]
    for color in required_colors:
        if color not in template["global"]["palette"]:
            raise ValueError(f"Template missing palette color: '{color}'")


def _csv_to_workbook(path):
    """Read a CSV file into a fresh openpyxl workbook.

    - Encoding: try utf-8-sig (handles BOM), fall back to gbk (common for
      Chinese Excel exports)
    - Separator: auto-detect comma vs tab
    - First row becomes the header row (detector handles the rest)

    Returns:
        An openpyxl Workbook with one sheet named after the file stem.
    """
    raw = path.read_bytes()

    # Decode with fallbacks
    text = None
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError(f"无法识别文件编码（已尝试 utf-8 / gbk）：{path}")

    # Auto-detect delimiter from the first line
    first_line = text.splitlines()[0] if text.splitlines() else ""
    delimiter = "\t" if "\t" in first_line and "," not in first_line else ","

    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows or all(not str(cell).strip() for row in rows for cell in row):
        raise ValueError(f"CSV 文件为空或无有效数据：{path}")

    # Drop trailing empty rows (trailing blank lines)
    while rows and all(not str(cell).strip() for cell in rows[-1]):
        rows.pop()

    wb = Workbook()
    ws = wb.active
    # Excel sheet names are limited to 31 chars
    ws.title = (path.stem[:31] or "Sheet1")
    for r, row in enumerate(rows, start=1):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)
    return wb


def process(
    input_path,
    output_path=None,
    theme="business_blue",
    template_path=None,
    watermark_text=None,
):
    """Process an Excel file: apply template styling and optional watermark.

    Args:
        input_path: Path to input .xlsx file
        output_path: Path to output .xlsx file. If None, overwrites input.
        theme: Built-in theme name (default: "business_blue")
        template_path: Custom template JSON path (overrides theme)
        watermark_text: Override watermark text from template

    Returns:
        Path to the output file
    """
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # Load template
    template = load_template(theme=theme, custom_path=template_path)

    # Override watermark text if provided; passing text also enables it
    if watermark_text is not None:
        wm = template.setdefault("watermark", {})
        wm["text"] = watermark_text
        if watermark_text:
            wm["enabled"] = True

    # Load workbook: .xlsx via openpyxl, .csv via built-in csv module
    if input_path.suffix.lower() == ".csv":
        wb = _csv_to_workbook(input_path)
    else:
        wb = load_workbook(input_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip empty sheets
        if ws.max_row is None or ws.max_row < 2:
            continue

        # Detect structure; None means no reliable header (cover/notes page) — skip
        structure = detect_structure(ws)
        if structure is None:
            continue

        # Detect column types
        col_types = detect_column_types(
            ws,
            structure["header_row"],
            structure["data_start"],
            structure["data_end"],
            structure["col_start"],
            structure["col_end"],
        )

        # Apply styles
        apply_all(ws, template, structure, col_types)

        # Apply watermark (to first sheet only)
        if sheet_name == wb.sheetnames[0]:
            apply_watermark(ws, template.get("watermark", {}), structure)

    # Save output
    if output_path is None:
        # CSV input has no style to preserve — write a styled .xlsx instead
        output_path = (
            input_path.with_suffix(".xlsx")
            if input_path.suffix.lower() == ".csv"
            else input_path
        )
    else:
        output_path = Path(output_path)
        if output_path.suffix.lower() != ".xlsx":
            raise ValueError(f"输出文件必须为 .xlsx 格式：{output_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return output_path
